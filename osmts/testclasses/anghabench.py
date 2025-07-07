from pathlib import Path
import fnmatch,tarfile,resource
import os,subprocess,shutil
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor

from .errors import GitCloneError


class AnghaBench:
    def __init__(self, **kwargs):
        self.rpms = set()
        self.believe_tmp:bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/AnghaBench')
        self.directory: Path = kwargs.get('saved_directory') / 'anghabench'
        self.log_files:Path = self.directory / 'log_files'
        self.matches:list = []
        self.failed_compile:list = []

        self.failed = 0
        self.total = 0

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'AnghaBench'

        self.ws.cell(1,1,"AnghaBench测试中编译未通过项目汇总")
        self.ws.merge_cells("A1:B1")
        self.ws.append(['c文件名','日志文件'])


    def pre_test(self):
        resource.setrlimit(resource.RLIMIT_NOFILE,(65536,524288))
        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)
        self.log_files.mkdir(parents=True)

        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            # 拉取AnghaBench的源码
            try:
                subprocess.run(
                    f"cd /root/osmts_tmp/ && git clone https://gitcode.com/qq_61653333/AnghaBench.git",
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE,
                )
            except subprocess.CalledProcessError as e:
                raise GitCloneError(e.returncode,'https://gitcode.com/qq_61653333/AnghaBench.git',e.stderr)



    def match2result(self,match):
        compile = subprocess.run(
            f"gcc {match[1]} -c -o {match[1]}.o",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT
        )
        if compile.returncode != 0: # 这里编译出错在预料之中不属于意外实践
            with open(self.log_files / (match[0] + '.log'), 'w') as log:
                log.write(compile.stdout.decode('utf-8'))
            return match[0]
        else:
            return None


    def run_test(self):
        if self.matches == []:
            for root,dirnames,filenames in os.walk(self.path):
                for filename in fnmatch.filter(filenames, '*.c'):
                    self.matches.append((filename,os.path.join(root,filename)))
        self.total = len(self.matches)
        with ThreadPoolExecutor(max_workers=os.cpu_count()) as executor:
            results = executor.map(self.match2result, self.matches)
            for result in results:
                if result is not None:
                    self.failed_compile.append(result)
        
        for item in set(self.failed_compile):
            self.ws.append([item, item + '.log'])

        self.failed = len(self.failed_compile)
        # 汇总结果
        self.ws.append([f"总共编译文件数{self.total}",f"通过编译数{self.total - self.failed}",f"失败编译数{self.failed}"])

        self.wb.save(self.directory / 'AnghaBench.xlsx')

        with tarfile.open(self.directory / 'AnghaBench.tar.xz',"w:xz") as tar:
            tar.add(self.log_files)


    def run(self):
        print('开始进行AnghaBench测试')
        self.pre_test()
        self.run_test()
        print('AnghaBench测试结束')