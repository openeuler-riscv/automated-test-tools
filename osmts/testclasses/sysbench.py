from openpyxl.workbook import Workbook
from pystemd.systemd1 import Unit
from pathlib import Path
import re,os
import pymysql,time
import sys,subprocess,shutil

from .errors import DefaultError,RunError


class sysBench:
    def __init__(self, **kwargs):
        self.rpms = {'sysbench','mysql-server'}
        self.directory: Path = kwargs.get('saved_directory') / 'sysbench'
        self.test_result:str = ''


    def pre_test(self):
        self.mysqld:Unit = Unit('mysqld.service',_autoload=True)
        try:
            self.mysqld.Unit.Start(b'replace')
        except:
            time.sleep(5)
            self.mysqld.load(force=True)
            self.mysqld.Unit.Start(b'replace')
        time.sleep(5)
        if self.mysqld.Unit.ActiveState != b'active':
            time.sleep(5)
            if self.mysqld.Unit.ActiveState != b'active':
                raise DefaultError(f"sysbench测试出错.开启mysqld.service失败,退出测试.")


        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)

        try:
            self.conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='',
            )
        except Exception as e:
            self.conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='123456',
            )
        cursor = self.conn.cursor()
        cursor.execute("ALTER USER 'root'@'localhost' IDENTIFIED BY '123456';")
        cursor.execute("DROP DATABASE IF EXISTS sysbench;")
        cursor.execute("CREATE DATABASE IF NOT EXISTS sysbench;")
        cursor.close()

        # 清理测试数据
        sysbench_clean = subprocess.run(
            "sysbench --db-driver=mysql --mysql-host=127.0.0.1 "
            "--mysql-port=3306 --mysql-user=root --mysql-password=123456 "
            "--mysql-db=sysbench --table_size=10000000 --tables=64 "
            f"--time=180 --threads={min(os.cpu_count(),16)} --report-interval=1 oltp_read_write cleanup",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
        )

        # 准备测试数据和表
        try:
            subprocess.run(
                "sysbench --db-driver=mysql --mysql-host=127.0.0.1 "
                "--mysql-port=3306 --mysql-user=root --mysql-password=123456 "
                "--mysql-db=sysbench --table_size=10000000 --tables=64 "
                f"--time=180 --threads={min(os.cpu_count(),16)} --report-interval=1 oltp_read_write prepare",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"sysbench测试出错.准备测试数据和表失败,报错信息:{e.stderr.decode('utf-8')}")


    def run_test(self):
        try:
            sysbench_run = subprocess.run(
                "sysbench --db-driver=mysql --mysql-host=127.0.0.1 "
                "--mysql-port=3306 --mysql-user=root --mysql-password=123456 "
                "--mysql-db=sysbench --table_size=10000000 --tables=64 "
                f"--time=180 --threads={min(os.cpu_count(),8)} " # --threads参数不能过大
                "--report-interval=1 oltp_read_write run",
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,e.stderr.decode('utf-8'))
        else:
            self.test_result = sysbench_run.stdout.decode('utf-8')
        with open(Path(self.directory) / 'sysbench.log', 'w') as log:
            log.write(self.test_result)


    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'sysbench'

        # SQL 统计
        ws.append(['SQL statistics[SQL统计]'])
        ws.merge_cells('A1:D1')
        ws.append(['标题','key','value','percent'])


        read_select = float(re.search(r"read:\s*(\d+)",self.test_result).group(1)) * 100
        write_select = float(re.search(r"write:\s*(\d+)",self.test_result).group(1)) * 100
        other_select = float(re.search(r"other:\s*(\d+)",self.test_result).group(1)) * 100
        total_select = float(re.search(r"total:\s*(\d+)",self.test_result).group(1))

        ws.append(['','读操作',read_select,"{:.4f}%".format(read_select / total_select)])
        ws.append(['','写操作',write_select,"{:.4f}%".format(write_select / total_select)])
        ws.append(['','其他操作',other_select,"{:.4f}%".format(other_select / total_select)])
        ws.append(['','总查询数量:',total_select,'/'])

        transactions = re.search(r"transactions:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        ws.append(['','总事务数:',transactions[0],'/'])
        ws.append(['','每秒事务数:',transactions[1],'/'])

        query_count = re.search(r"queries:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        ws.append(['','总查询数:',query_count[0],'/'])
        ws.append(['','每秒查询数:',query_count[1],'/'])

        ignore_errors = re.search(r"ignored errors:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        ws.append(['','忽略错误数:',ignore_errors[0],'/'])
        ws.append(['','每秒忽略错误数:',ignore_errors[1],'/'])

        reconnects = re.search(r"reconnects:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        ws.append(['','重连次数:',reconnects[0],'/'])
        ws.append(['','每秒重连次数:',reconnects[1],'/'])

        ws.cell(3, 1, "执行的查询")
        ws.merge_cells('A3:A15')


        # 通用统计
        ws.append(['General statistics[通用统计]'])
        ws.merge_cells('A16:D16')
        total_time = re.search(r"total time:\s*(\d+\.\d+)s",self.test_result).group(1)
        total_number_of_events = re.search(r"total number of events:\s*(\d+)",self.test_result).group(1)
        ws.append(['','测试总时间:',total_time + 's','/'])
        ws.append(['','总事务数:',total_number_of_events,'/'])
        ws.merge_cells('A17:A18')


        # 延迟统计
        ws.append(['Latency(ms)[延迟统计]'])
        min = re.search(r"min:\s*(\d+\.\d+)",self.test_result).group(1)
        avg = re.search(r"avg:\s*(\d+\.\d+)",self.test_result).group(1)
        max = re.search(r"max:\s*(\d+\.\d+)",self.test_result).group(1)
        percentile_95th = re.search(r"95th percentile:\s*(\d+\.\d+)",self.test_result).group(1)
        sum = re.search(r"sum:\s*(\d+\.\d+)",self.test_result).group(1)

        ws.append(['','最小延迟:',min,'/'])
        ws.append(['', '平均延迟:', avg, '/'])
        ws.append(['', '最大延迟:', max, '/'])
        ws.append(['', '95% 的查询延迟小于:', percentile_95th, '/'])
        ws.append(['', '总延迟时间:', sum, '/'])

        ws.merge_cells("A20:A24")

        # 线程公平性
        ws.append(['Threads fairness[线程公平性]'])
        ws.merge_cells("A25:D25")

        events_avg,events_stddev = re.search(r"events \(avg/stddev\):\s*(\d+\.\d+)/(\d+\.\d+)",self.test_result).groups()
        ws.append(['','每个线程平均处理事件数:',events_avg,'/'])
        ws.append(['', '每个线程平均处理标准差:', events_stddev, '越小越好,负载均衡'])

        execution_time_avg,execution_time_stddev = re.search(r"execution time \(avg/stddev\):\s*(\d+\.\d+)/(\d+\.\d+)",self.test_result).groups()
        ws.append(['','每个线程平均执行时间:',execution_time_avg + 's','/'])
        ws.append(['','每个线程平均执行时间标准差:',execution_time_stddev,'越小越好,说明线程执行时间非常均匀'])

        ws.merge_cells("A26:A29")

        wb.save(self.directory / 'sysbench.xlsx')



    def post_test(self):
        # 清理测试数据
        sysbench_clean = subprocess.run(
            "sysbench --db-driver=mysql --mysql-host=127.0.0.1 "
            "--mysql-port=3306 --mysql-user=root --mysql-password=123456 "
            "--mysql-db=sysbench --table_size=10000000 --tables=64 "
            f"--time=180 --threads={min(os.cpu_count(),16)} --report-interval=1 oltp_read_write cleanup",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
        )
        if sysbench_clean.returncode != 0:
            print(f"sysbench测试.清理测试数据失败,报错信息:{sysbench_clean.stderr.decode('utf-8')}")

        self.mysqld.Unit.Stop(b'replace')
        subprocess.run(
            "dnf remove -y mysql-server",
            shell=True,stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL
        )


    def run(self):
        print('开始进行sysbench测试')
        self.pre_test()
        self.run_test()
        self.result2summary()
        self.post_test()
        print('sysbench测试结束')