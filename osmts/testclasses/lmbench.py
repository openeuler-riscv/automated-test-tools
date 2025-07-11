from pathlib import Path
import subprocess,shutil,pexpect
from openpyxl import Workbook

from .errors import GitCloneError,RunError,SummaryError


class Lmbench:
    def __init__(self, **kwargs):
        self.rpms = {'libtirpc-devel'}
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/lmbench')
        self.directory: Path = kwargs.get('saved_directory') / 'lmbench'
        self.compiler: str = kwargs.get('compiler')
        self.test_result = ''


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True, parents=True)
        if self.path.exists():
            if self.path.is_file():
                self.path.unlink()
            else:
                shutil.rmtree(self.path)

        # 获取lmbench源码
        try:
            subprocess.run(
                "git clone https://gitee.com/April_Zhao/lmbench.git",
                cwd="/root/osmts_tmp",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise GitCloneError(e.returncode,'https://gitee.com/April_Zhao/lmbench.git',e.stderr.decode())


    def run_test(self):
        # make后直接就运行了
        make = pexpect.spawn(
            command = '/bin/bash',
            args = ['-c',f"make CC={self.compiler} results"],
            cwd=self.path,
            encoding = 'utf-8',
            logfile = open(self.directory / 'osmts_lmbench.log', 'w')
        )

        # 同时运行lmbench的份数
        make.expect_exact("MULTIPLE COPIES [default 1]:",timeout=180)
        make.sendline("1")

        # 允许作业调度
        make.expect_exact("Job placement selection [default 1]:",timeout=180)
        make.sendline("1")

        # 设置测试内存大小
        make.expect_exact("MB [default",timeout=180)
        make.sendline("4096")

        # 选择要运行的测试集
        make.expect_exact("SUBSET (ALL|HARWARE|OS|DEVELOPMENT) [default all]:",timeout=1800)
        make.sendline("ALL")

        # 不跳过内存latency测试
        make.expect_exact('FASTMEM [default no]:',timeout=1800)
        make.sendline("no")

        # 不跳过文件系统latency测试
        make.expect_exact('SLOWFS [default no]:',timeout=1800)
        make.sendline("no")

        # 不测试硬盘
        make.expect_exact('DISKS [default none]:',timeout=1800)
        make.sendline()

        # 不测试网络
        make.expect_exact("REMOTE [default none]:",timeout=1800)
        make.sendline("")

        # 测试CPU与设定频率
        make.expect_exact('Processor mhz',timeout=1800)
        make.sendline()

        # 设定临时目录存放测试文件
        make.expect_exact('FSDIR [default /usr/tmp]:',timeout=1800)
        make.sendline('/usr/tmp')

        # 设置测试输出信息文件存放目录
        make.expect_exact('Status output file [default /dev/tty]:',timeout=1800)
        make.sendline('/dev/tty')

        # 设置不发邮件
        make.expect_exact('Mail results [default yes]:',timeout=1800)
        make.sendline('no')

        # 等待lmbench测试运行结束
        make.expect_exact(pexpect.EOF,timeout=18000)


        # 获取运行结果
        try:
            subprocess.run(
                "cd /root/osmts_tmp/lmbench && make see",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stdin=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,e.stderr.decode('utf-8'))

        shutil.copyfile('/root/osmts_tmp/lmbench/results/summary.out',self.directory / 'lmbench_summary.out')


    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'lmbench'
        summary_out = open('/root/osmts_tmp/lmbench/results/summary.out','r')
        lines = summary_out.readlines()
        summary_out.close()
        ws.cell(1,1,'Basic system parameters')
        ws.merge_cells('A1:A8')
        ws.cell(1,2,'Host')
        ws.cell(2,2,'OS')
        ws.cell(3,2,'Description')
        ws.cell(4,2,'Mhz')
        ws.cell(5,2,'tlb pages')
        ws.cell(6,2,'cache line')
        ws.cell(7,2,'mem par')
        ws.cell(8,2,'scal load')
        content = [line for line in lines[12].split(' ') if line != '']
        ws.cell(1,3,content[0])
        ws.cell(2,3,content[1] + content[2])
        ws.cell(3,3,content[3])
        ws.cell(4,3,content[4])
        ws.cell(6,3,content[5])
        ws.cell(7,3,content[6])
        try:
            ws.cell(8,3,content[7].rstrip('\n'))
        except IndexError:
            ws.cell(8,3,'Empty')

        ws.cell(9,1,'Processor, Processes - times in microseconds - smaller is better')
        ws.merge_cells('A9:A21')
        ws.cell(9,2,'Host')
        ws.cell(10, 2, 'OS')
        ws.cell(11, 2, 'Mhz')
        ws.cell(12, 2, 'null call')
        ws.cell(13, 2, 'null I/O')
        ws.cell(14, 2, 'stat')
        ws.cell(15, 2, 'open clos')
        ws.cell(16, 2, 'slct TCP')
        ws.cell(17, 2, 'sig inst')
        ws.cell(18, 2, 'sig hndl')
        ws.cell(19, 2, 'fork proc')
        ws.cell(20, 2, 'exec proc')
        ws.cell(21, 2, 'sh proc')
        content = [line for line in lines[19].split(' ') if line != '']
        ws.cell(9,3,content[0])
        ws.cell(10, 3, content[1] + content[2])
        ws.cell(11,3,content[3])
        ws.cell(12, 3, content[4])
        ws.cell(13, 3, content[5])
        ws.cell(14, 3, content[6])
        ws.cell(15, 3, content[7])
        ws.cell(16, 3, content[8])
        ws.cell(17, 3, content[9])
        ws.cell(18, 3, content[10])
        ws.cell(19, 3, content[11])
        ws.cell(20, 3, content[12])
        try:
            ws.cell(21, 3, content[13].rstrip('\n'))
        except IndexError:
            ws.cell(21,3,'Empty')

        ws.cell(22,1,'Basic integer operations - times in nanoseconds - smaller is better')
        ws.merge_cells("A22:A28")
        ws.cell(22,2,'Host')
        ws.cell(23, 2, 'OS')
        ws.cell(24, 2, 'intgr bit')
        ws.cell(25, 2, 'intgr add')
        ws.cell(26, 2, 'intgr mul')
        ws.cell(27, 2, 'intgr div')
        ws.cell(28, 2, 'intgr mod')
        content = [line for line in lines[26].split(' ') if line != '']
        ws.cell(22, 3, content[0])
        ws.cell(23, 3, content[1] + content[2])
        ws.cell(24, 3, content[3])
        ws.cell(25, 3, content[4])
        ws.cell(26, 3, content[5])
        ws.cell(27, 3, content[6])
        try:
            ws.cell(28, 3, content[7].rstrip('\n'))
        except IndexError:
            ws.cell(28,3,'Empty')

        ws.cell(29,1,'Basic uint64 operations - times in nanoseconds - smaller is better')
        ws.merge_cells("A29:A35")
        ws.cell(29,2,'Host')
        ws.cell(30, 2, 'OS')
        ws.cell(31, 2, 'int64 bit')
        ws.cell(32, 2, 'int64 add')
        ws.cell(33, 2, 'int64 mul')
        ws.cell(34, 2, 'int64 div')
        ws.cell(35, 2, 'int64 mod')
        content = [line for line in lines[33].split(' ') if line != '']
        ws.cell(29,3,content[0])
        ws.cell(30, 3, content[1] + content[2])
        ws.cell(31, 3, content[3])
        ws.cell(32, 3, '')
        ws.cell(33, 3, content[4])
        ws.cell(34, 3, content[5])
        try:
            ws.cell(35, 3, content[6].rstrip('\n'))
        except IndexError:
            ws.cell(35,3,'Empty')

        ws.cell(36,1,'Basic float operations - times in nanoseconds - smaller is better')
        ws.merge_cells("A36:A41")
        ws.cell(36,2,'Host')
        ws.cell(37, 2, 'OS')
        ws.cell(38, 2, 'float bit')
        ws.cell(39, 2, 'float add')
        ws.cell(40, 2, 'float mul')
        ws.cell(41, 2, 'float bogo')
        content = [line for line in lines[40].split(' ') if line != '']
        ws.cell(36,3,content[0])
        ws.cell(37, 3, content[1] + content[2])
        ws.cell(38, 3, content[3])
        ws.cell(39, 3, content[4])
        ws.cell(40, 3, content[5])
        try:
            ws.cell(41, 3, content[6].rstrip('\n'))
        except IndexError:
            ws.cell(41,3,'Empty')

        ws.cell(42, 1, 'Basic double operations - times in nanoseconds - smaller is better')
        ws.merge_cells("A42:A45")
        ws.cell(42,2,'Host')
        ws.cell(43, 2, 'OS')
        ws.cell(44, 2, 'double bit')
        ws.cell(45, 2, 'double add')
        ws.cell(46, 2, 'double mul')
        ws.cell(47, 2, 'double bogo')
        content = [line for line in lines[47].split(' ') if line != '']
        ws.cell(42,3,content[0])
        ws.cell(43, 3, content[1] + content[2])
        ws.cell(44, 3, content[3])
        ws.cell(45, 3, content[4])
        ws.cell(46, 3, content[5])
        try:
            ws.cell(47, 3, content[6].rstrip('\n'))
        except IndexError:
            ws.cell(47,3,'Empty')

        ws.cell(48, 1,'Context switching - times in microseconds - smaller is better')
        ws.merge_cells("A48:A56")
        ws.cell(48,2,'Host')
        ws.cell(49,2,'OS')
        ws.cell(50,2,'2p/0K|ctxsw')
        ws.cell(51,2,'2p/16K|ctxsw')
        ws.cell(52,2,'2p/64K|ctxsw')
        ws.cell(53,2,'8p/16K|ctxsw')
        ws.cell(54,2,'8p/64K|ctxsw')
        ws.cell(55,2,'16p/16K|ctxsw')
        ws.cell(56,2,'16p/64K|ctxsw')
        content = [line for line in lines[54].split(' ') if line != '']
        ws.cell(48,3,content[0])
        ws.cell(49,3,content[1] + content[2])
        ws.cell(50,3,content[3])
        ws.cell(51,3,content[4])
        ws.cell(52,3,content[5])
        ws.cell(53,3,content[6])
        ws.cell(54,3,content[7])
        ws.cell(55,3,content[8])
        try:
            ws.cell(56,3,content[9].rstrip('\n'))
        except IndexError:
            ws.cell(56,3,'Empty')

        ws.cell(57,1,'*Local* Communication latencies in microseconds - smaller is better')
        ws.merge_cells("A57:A66")
        ws.cell(57,2,'Host')
        ws.cell(58,2,'OS')
        ws.cell(59,2,'2p/0K|ctxsw')
        ws.cell(60,2,'PIPE')
        ws.cell(61,2,'AF UNIX')
        ws.cell(62,2,'UDP')
        ws.cell(63,2,'RPC/UDP')
        ws.cell(64,2,'TCP')
        ws.cell(65,2,'RPC/TCP')
        ws.cell(66,2,'TCP conn')
        content = [line for line in lines[61].split(' ') if line != '']
        ws.cell(57,3,content[0])
        ws.cell(58,3,content[1] + content[2])
        ws.cell(59,3,content[3])
        ws.cell(60,3,content[4])
        ws.cell(61,3,content[5])
        ws.cell(62,3,content[6])
        ws.cell(63,3,content[7])
        ws.cell(64,3,content[8])
        ws.cell(65,3,content[9])
        try:
            ws.cell(66,3,content[10].rstrip('\n'))
        except IndexError:
            ws.cell(66,3,'Empty')

        # 网络测试前面选择不测,故此少了一组

        ws.cell(67,1,'File & VM system latencies in microseconds - smaller is better')
        ws.merge_cells("A67:A76")
        ws.cell(67, 2, 'Host')
        ws.cell(68, 2, 'OS')
        ws.cell(69, 2, '0K File Create')
        ws.cell(70, 2, '0K File Delete')
        ws.cell(71, 2, '10K File Create')
        ws.cell(72, 2, '10K File Create')
        ws.cell(73, 2, 'Mmap Latency')
        ws.cell(74, 2, 'Prot Fault')
        ws.cell(75, 2, 'Page Fault')
        ws.cell(76, 2, '100fd selct')
        content = [line for line in lines[75].split(' ') if line != '']
        ws.cell(67, 3, content[0])
        ws.cell(68, 3, content[1] + content[2])
        ws.cell(69, 3, content[3])
        ws.cell(70, 3, content[4])
        ws.cell(71, 3, content[5])
        ws.cell(72, 3, content[6])
        ws.cell(73, 3, '')
        ws.cell(74, 3, content[7])
        ws.cell(75, 3, '')
        try:
            ws.cell(76, 3, content[8].rstrip('\n'))
        except IndexError:
            ws.cell(76,3,'Empty')

        ws.cell(77, 1, '*Local* Communication bandwidths in MB/s - bigger is better')
        ws.merge_cells("A77:A87")
        ws.cell(77,2,'Host')
        ws.cell(78,2,'OS')
        ws.cell(79,2,'PIPE')
        ws.cell(80,2,'AF UNIX')
        ws.cell(81,2,'TCP')
        ws.cell(82,2,'File reread')
        ws.cell(83,2,'Mmap reread')
        ws.cell(84,2,'Bcopy(libc)')
        ws.cell(85,2,'Bcopy(hand)')
        ws.cell(86,2,'Mem read')
        ws.cell(87,2,'Mem write')
        content = [line for line in lines[82].split(' ') if line != '']
        ws.cell(77,3,content[0])
        ws.cell(78,3,content[1] + content[2])
        ws.cell(79,3,content[3])
        ws.cell(80,3,content[4])
        ws.cell(81,3,content[5])
        ws.cell(82,3,content[6])
        ws.cell(83,3,content[7])
        ws.cell(84,3,content[8])
        ws.cell(85,3,content[9])
        ws.cell(86,3,content[10])
        try:
            ws.cell(87,3,content[11].rstrip('\n'))
        except IndexError:
            ws.cell(87,3,'Empty')

        ws.cell(88,1,'Memory latencies in nanoseconds - smaller is better\n(WARNING - may not be correct, check graphs)')
        ws.merge_cells("A88:A95")
        ws.cell(88,2,'Host')
        ws.cell(89,2,'OS')
        ws.cell(90,2,'Mhz')
        ws.cell(91,2,'L1 $')
        ws.cell(92,2,'L2 $')
        ws.cell(93,2,'Main mem')
        ws.cell(94,2,'Rand mem')
        ws.cell(95,2,'Guesses')
        content = [line for line in lines[89].split(' ') if line != '']
        ws.cell(88,3,content[0])
        ws.cell(89,3,content[1] + content[2])
        ws.cell(90,3,content[3])
        ws.cell(91,3,content[4])
        ws.cell(92,3,content[5])
        ws.cell(93,3,content[6])
        try:
            ws.cell(94,3,content[7].rstrip('\n'))
        except IndexError:
            ws.cell(94,3,'Empty')
        ws.cell(95,3,'')


        wb.save(self.directory / 'lmbench.xlsx')


    def run(self):
        print("开始进行lmbench测试")
        self.pre_test()
        self.run_test()
        try:
            self.result2summary()
        except Exception as e:
            logFile = self.directory / 'lmbench_summary_error.log'
            with open(logFile, 'w') as log:
                log.write(str(e))
            raise SummaryError(logFile)
        print("lmbench测试结束")