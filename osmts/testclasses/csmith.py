import os
from pathlib import Path
import sys,subprocess,shutil
from openpyxl import Workbook
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor,as_completed

from .errors import GitCloneError,DefaultError


class Csmith:
    def __init__(self, **kwargs):
        self.rpms = {'g++','m4'}
        self.path = Path('/root/osmts_tmp/csmith')
        self.directory: Path = kwargs.get('saved_directory') / 'csmith'
        self.source:Path = self.directory / 'source'
        self.bin: Path = self.directory / 'bin'
        self.csmith_count:int = kwargs.get('csmith_count',1000) + 1


    def create_source_and_bin(self,number):
        source_code = f"{self.source}/csmith{number}.c"
        include_directory = self.path / "install" / "include"
        # 创建随机c文件
        csmith = subprocess.run(
            "/root/osmts_tmp/csmith/install/bin/csmith",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        if csmith.returncode != 0:
            print(f"csmith测试出错.生成c代码失败,报错信息:{csmith.stderr.decode('utf-8')}")
            sys.exit(1)
        with open(f"{self.source}/csmith{number}.c", "w") as file:
            file.write(csmith.stdout.decode('utf-8'))

        # 分别用gcc和clang编译c文件
        compile = subprocess.run(
            f"gcc {source_code} -I {include_directory} -o {self.bin}/csmith{number}_gcc && "
            f"clang {source_code} -I {include_directory} -o {self.bin}/csmith{number}_clang",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        if compile.returncode != 0:
            print(f"csmith测试出错.编译c代码失败,报错信息:{compile.stderr.decode('utf-8')}")
            sys.exit(1)


    def pre_test(self):
        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)
        if not self.source.exists():
            self.source.mkdir()
        if not self.bin.exists():
            self.bin.mkdir()

        if self.path.exists():
            shutil.rmtree(self.path)
        self.path.mkdir(parents=True,exist_ok=True)

        try:
            subprocess.run(
                "git clone https://gitcode.com/qq_61653333/csmith.git -b master",
                cwd="/root/osmts_tmp/",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise GitCloneError(e.returncode,'https://gitcode.com/qq_61653333/csmith.git',e.stderr.decode('utf-8'))


        try:
            subprocess.run(
                f"mkdir install && cmake -DCMAKE_INSTALL_PREFIX=/root/osmts_tmp/csmith/install . && make -j {os.cpu_count()} && make install",
                cwd=self.path,
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"Csmith测试出错.构建csmith项目失败,报错信息:{e.stderr.decode('utf-8')}")


        # 批量生成c代码
        with ThreadPoolExecutor(max_workers=os.cpu_count()) as pool:
            for _ in tqdm(pool.map(self.create_source_and_bin, range(1,self.csmith_count)),desc='批量生成c代码进度',total=self.csmith_count):
                pass

        print(f'源码文件生成在{self.source}目录,已完成')
        print(f'二进制文件生成在{self.bin}目录,已完成')


    def check_each_csmith(self,id:int) -> tuple:
        # 超过10秒则跳过(生成的c代码要求算力太大,不符合测试条件)
        gcc = subprocess.run(
            f"timeout 10 {self.directory}/bin/csmith{id}_gcc",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL
        )
        clang = subprocess.run(
            f"timeout 10 {self.directory}/bin/csmith{id}_clang",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL
        )
        if gcc.returncode != 0 or clang.returncode != 0:
            return (id,None,None)
        return (id,gcc.stdout.decode('utf-8'), clang.stdout.decode('utf-8'))


    def run_test(self):
        with ThreadPoolExecutor(max_workers=os.cpu_count()) as pool:
            # 先提交任务
            futures = [pool.submit(self.check_each_csmith,i) for i in range(1,self.csmith_count)]

            wb = Workbook()
            ws = wb.active
            ws.title = 'Csmith'
            ws.cell(1, 1, "程序名")
            ws.cell(1, 2, "检验和是否一致")
            ws.cell(1, 3, "gcc checksum")
            ws.cell(1, 4, "clang checksum")
            for i in range(1, self.csmith_count):
                ws.cell(i + 1, 1, f"csmith{i}.c")

            # 获取返回值
            for future in tqdm(as_completed(futures),total=self.csmith_count,desc="csmith完成进度"):
                line,gcc_checksum,clang_checksum = future.result()
                line += 1
                if gcc_checksum is None and clang_checksum is None:
                    ws.cell(line, 2, "程序运行超时,不符合条件")
                elif gcc_checksum == clang_checksum:
                    ws.cell(line, 2, "是")
                else:
                    ws.cell(line, 2, "否")
                    ws.cell(line, 3, gcc_checksum)
                    ws.cell(line, 4, clang_checksum)

            wb.save(self.directory / 'csmith.xlsx')


    def run(self):
        print('开始进行Csmith测试')
        self.pre_test()
        self.run_test()
        print('Csmith测试结束')
