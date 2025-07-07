from pathlib import Path
import re,subprocess,shutil
from openpyxl import Workbook
from .errors import GitCloneError,CompileError,SummaryError,RunError



class Stream:
    def __init__(self,**kwargs ):
        self.rpms = set()
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/stream')
        self.directory:Path = kwargs.get('saved_directory') / 'stream'
        self.compiler:str = kwargs.get('compiler')
        self.test_result = ''


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)

        # believe_tmp表明尽可能相信/root/osmts_tmp/
        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            try:
                subprocess.run(
                    "git clone https://gitee.com/April_Zhao/stream.git",
                    cwd="/root/osmts_tmp",
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE,
                )
            except subprocess.CalledProcessError as e:
                raise GitCloneError(
                    error_code = e.returncode,
                    url = 'https://gitee.com/April_Zhao/stream.git',
                    stderr = e.stderr.decode('utf-8'),
                )


        # 编译stream.c
        try:
            subprocess.run(
                f"{self.compiler} -O3 -fopenmp -DSTREAM_ARRAY_SIZE=35000000 -DNTIMES=50 stream.c -o stream_O3",
                cwd=self.path,
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise CompileError(
                error_code = e.returncode,
                compiler = self.compiler,
                stderr = e.stderr.decode('utf-8')
            )


    def run_test(self):
        try:
            stream_o3 = subprocess.run(
                f"./stream_O3",
                cwd=self.path,
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            log_file = self.directory / 'run_stream_O3_failed.log'
            with open(log_file,'w') as log:
                log.write(e.stderr.decode('utf-8'))
            raise RunError(e.returncode, f"详细信息请查看{log_file}日志文件")
        else:
            # 运行成功,保存日志
            self.test_result = stream_o3.stdout.decode('utf-8')
            with open(self.path / 'stream.log','w',encoding='utf-8') as file:
                file.write(self.test_result)


    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "stream"
        ws.cell(1, 1, "Function")
        ws.cell(1, 2, "Best Rate MB/s")
        ws.cell(1, 3, "Avg time")
        ws.cell(1, 4, "Min time")
        ws.cell(1, 5, "Max time")
        ws.cell(2, 1, "Copy")
        ws.cell(3, 1, "Scale")
        ws.cell(4, 1, "Add")
        ws.cell(5, 1, "Triad")

        col_list = ["Copy:", "Scale:", "Add:", "Triad:"]
        for i in range(2, 6):
            Function = re.search(rf'^\s*{col_list[i - 2]}\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)',
                                 self.test_result,
                                 re.MULTILINE | re.IGNORECASE)
            ws.cell(i, 1, col_list[i - 2])
            for j in range(2, 6):
                ws.cell(i, j, Function.group(j - 1))
        # 保存
        wb.save(self.directory / 'stream.xlsx')



    def run(self):
        print("开始进行stream测试")
        self.pre_test()
        self.run_test()
        try:
            self.result2summary()
        except Exception as e:
            fileName = self.directory / 'summary_stream_error.log'
            with open(fileName,'w') as log:
                log.write(str(e))
            raise SummaryError(fileName)
        print("stream测试结束")
