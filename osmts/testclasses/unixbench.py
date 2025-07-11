
import subprocess, fileinput, shutil, re, os
from pathlib import Path
from openpyxl import Workbook
from .errors import GitCloneError, CompileError, RunError, SummaryError

class Unixbench:
    def __init__(self, **kwargs):
        self.rpms = {'perl', 'perl-CPAN'}
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/byte-unixbench')
        self.directory: Path = kwargs.get('saved_directory') / 'unixbench'
        self.compiler: str = kwargs.get('compiler')
        self.test_results = []
        self.cpu_count = os.cpu_count()

    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True, parents=True)
        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            try:
                subprocess.run(
                    args="git clone https://gitcode.com/gh_mirrors/by/byte-unixbench.git",
                    cwd="/root/osmts_tmp",
                    shell=True, check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise GitCloneError(e.returncode, 'https://gitcode.com/gh_mirrors/by/byte-unixbench.git', e.stderr)

        if self.compiler == 'clang':
            for line in fileinput.input('/root/osmts_tmp/byte-unixbench/UnixBench/Makefile', inplace=True):
                if 'CC=gcc' in line:
                    line = line.replace('CC=gcc', 'CC=clang')
                print(line, end='')

        try:
            subprocess.run(
                "make",
                cwd="/root/osmts_tmp/byte-unixbench/UnixBench",
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise CompileError(e.returncode, self.compiler, e.stderr)

    def run_test(self, times=3):
        for i in range(1, times + 1):
            try:
                run = subprocess.run(
                    "./Run -c 1 -c $(nproc)",
                    cwd="/root/osmts_tmp/byte-unixbench/UnixBench/",
                    shell=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise RunError(e.returncode, e.stderr)
            else:
                result = run.stdout.decode('utf-8')
                self.test_results.append(result)
                log_path = self.directory / f'unixbench_{i}.log'
                with open(log_path, 'w') as file:
                    file.write(result)

    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "unixbench"

        ws['a1'] = "unixbench"
        ws['b1'] = "测试项目"
        ws['c1'] = "Raw Results Avg(原始数据)"
        ws['d1'] = "Index Values Avg(基准值)"

        ws['a2'] = f"{self.cpu_count} CPUs & 1 parallel"
        ws['a16'] = f"{self.cpu_count} CPUs & {self.cpu_count} parallel"
        ws.merge_cells("a2:a14")
        ws.merge_cells("a16:a28")

        test_items = [
            "Dhrystone 2 using register variables(lps)",
            "Double-Precision Whetstone(MWIPS)",
            "Execl Throughput(lps)",
            "File Copy 1024 bufsize 2000 maxblocks(KBps)",
            "File Copy 256 bufsize 500 maxblocks(KBps)",
            "File Copy 4096 bufsize 8000 maxblocks(KBps)",
            "Pipe Throughput(lps)",
            "Pipe-based Context Switching(lps)",
            "Process Creation(lps)",
            "Shell Scripts (1 concurrent)(lpm)",
            "Shell Scripts (8 concurrent)(lpm)",
            "System Call Overhead(lps)",
            "System Benchmarks Index Score"
        ]

        for i in range(13):
            ws[f'b{2 + i}'] = test_items[i]
            ws[f'b{16 + i}'] = test_items[i]

        match_patterns = [
            re.compile(r"Dhrystone 2 using register variables\s+([\d.]+)\s+lps"),
            re.compile(r"Double-Precision Whetstone\s+([\d.]+)\s+MWIPS"),
            re.compile(r"Execl Throughput\s+([\d.]+)\s+lps"),
            re.compile(r"File Copy 1024 bufsize 2000 maxblocks\s+([\d.]+)\s+KBps"),
            re.compile(r"File Copy 256 bufsize 500 maxblocks\s+([\d.]+)\s+KBps"),
            re.compile(r"File Copy 4096 bufsize 8000 maxblocks\s+([\d.]+)\s+KBps"),
            re.compile(r"Pipe Throughput\s+([\d.]+)\s+lps"),
            re.compile(r"Pipe-based Context Switching\s+([\d.]+)\s+lps"),
            re.compile(r"Process Creation\s+([\d.]+)\s+lps"),
            re.compile(r"Shell Scripts \(1 concurrent\)\s+([\d.]+)\s+lpm"),
            re.compile(r"Shell Scripts \(8 concurrent\)\s+([\d.]+)\s+lpm"),
            re.compile(r"System Call Overhead\s+([\d.]+)\s+lps"),
        ]

        raw_data_1p = [[] for _ in range(13)]
        raw_data_np = [[] for _ in range(13)]

        for content in self.test_results:
            for i, pattern in enumerate(match_patterns):
                matches = pattern.findall(content)
                if len(matches) >= 2:
                    raw_data_1p[i].append(float(matches[0]))
                    raw_data_np[i].append(float(matches[1]))
                else:
                    raise SummaryError(f"Missing match for {test_items[i]}")

        for i in range(13):
            if raw_data_1p[i]:
                ws[f'c{2 + i}'] = round(sum(raw_data_1p[i]) / len(raw_data_1p[i]), 2)
            if raw_data_np[i]:
                ws[f'c{16 + i}'] = round(sum(raw_data_np[i]) / len(raw_data_np[i]), 2)

        index_scores_1p = [[] for _ in range(13)]
        index_scores_np = [[] for _ in range(13)]

        for content in self.test_results:
            scores = re.findall(r"\b(\d+\.\d+)\s*$", content, re.MULTILINE)
            if len(scores) >= 26:
                for i in range(13):
                    index_scores_1p[i].append(float(scores[i]))
                    index_scores_np[i].append(float(scores[i + 13]))
            else:
                raise SummaryError("Index score count is insufficient")

        for row in range(13):
            if index_scores_1p[row]:
                ws[f'd{2 + row}'] = round(sum(index_scores_1p[row]) / len(index_scores_1p[row]), 2)
            if index_scores_np[row]:
                ws[f'd{16 + row}'] = round(sum(index_scores_np[row]) / len(index_scores_np[row]), 2)

        wb.save(self.directory / 'unixbench.xlsx')

    def run(self):
        print("开始进行unixbench测试")
        self.pre_test()
        self.run_test(times=3)
        try:
            self.result2summary()
        except Exception as e:
            logFile = self.directory / 'unixbench_summary_error.log'
            with open(logFile, 'w') as log:
                log.write(str(e))
            raise SummaryError(logFile)
        print("unixbench测试结束")
