import shutil
import subprocess
import os,threading
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from openpyxl import Workbook

from .errors import GitCloneError,RunError,DefaultError


class Jotai:
    def __init__(self, **kwargs):
        self.rpms = set()
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/jotai-benchmarks')
        self.directory: Path = kwargs.get('saved_directory') / 'jotai'

        self.anghaLeaves = Path('/root/osmts_tmp/jotai-benchmarks/benchmarks/anghaLeaves')
        self.anghaLeaves_output = self.directory / 'anghaLeaves_output'
        self.anghaLeaves_logs = self.directory / 'anghaLeaves_logs'

        self.anghaMath = Path('/root/osmts_tmp/jotai-benchmarks/benchmarks/anghaMath')
        self.anghaMath_output = self.directory / 'anghaMath_output'
        self.anghaMath_logs = self.directory / 'anghaMath_logs'

        self.wb = Workbook()

        # anghaLeaves 数据表
        self.ws_anghaLeaves = self.wb.active
        self.ws_anghaLeaves.title = 'Jotai-anghaLeaves'
        self.ws_anghaLeaves.cell(1,1,"c文件名")
        self.ws_anghaLeaves.cell(1, 2, "编译/运行是否通过")
        self.ws_anghaLeaves.cell(1, 3, "日志文件")
        self.anghaLeaves_passed = 0
        self.anghaLeaves_failed = 0
        self.anghaLeaves_lock = threading.Lock()

        # anghaMath 数据表
        self.ws_anghaMath = self.wb.create_sheet(title='Jotai-anghaMath')
        self.ws_anghaMath.cell(1,1,"c文件名")
        self.ws_anghaMath.cell(1, 2, "编译/运行是否通过")
        self.ws_anghaMath.cell(1, 3, "日志文件")
        self.anghaMath_passed = 0
        self.anghaMath_failed = 0
        self.anghaMath_lock = threading.Lock()


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(parents=True)
        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            # 拉取jotai-benchmarks源码
            try:
                subprocess.run(
                    "git clone https://gitcode.com/qq_61653333/jotai-benchmarks.git",
                    cwd="/root/osmts_tmp",
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE,
                )
            except subprocess.CalledProcessError as e:
                raise GitCloneError(e.returncode,'https://gitcode.com/qq_61653333/jotai-benchmarks.git',e.stderr.decode())


        # 创建anghaLeaves系列目录
        if self.anghaLeaves_output.exists():
            shutil.rmtree(self.anghaLeaves_output)
        if self.anghaLeaves_logs.exists():
            shutil.rmtree(self.anghaLeaves_logs)
        self.anghaLeaves_output.mkdir(parents=True)
        self.anghaLeaves_logs.mkdir(parents=True)

        # 创建anghaMath系列目录
        if self.anghaMath_output.exists():
            shutil.rmtree(self.anghaMath_output)
        if self.anghaMath_logs.exists():
            shutil.rmtree(self.anghaMath_logs)
        self.anghaMath_output.mkdir(parents=True)
        self.anghaMath_logs.mkdir(parents=True)


    # 编译并运行anghaLeaves程序组
    def compile_and_run_anghaLeaves(self,source_file):
        complete_source_file = self.anghaLeaves / source_file
        log_file = f"{self.anghaLeaves_logs}/{source_file}.log"

        # 编译
        compile = subprocess.run(
            f"gcc {complete_source_file} -o {self.anghaLeaves_output}/{source_file}.out",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
        )
        if compile.returncode != 0:
            with self.anghaLeaves_lock:
                self.anghaLeaves_failed += 1
                self.ws_anghaLeaves.append([source_file, 'compile failed',log_file])
            with open(log_file,'w') as log:
                log.write(compile.stdout.decode('utf-8'))
            return # 编译失败就不再往下运行了

        # 运行
        run = subprocess.run(
            f"{self.anghaLeaves_output}/{source_file}.out 0",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
        )
        if run.returncode != 0:
            with self.anghaLeaves_lock:
                self.anghaLeaves_failed += 1
                self.ws_anghaLeaves.append([source_file, 'run failed',log_file])
            with open(log_file,'w') as log:
                log.write(compile.stdout.decode('utf-8') + '\n' + run.stdout.decode('utf-8'))
        else:
            with self.anghaLeaves_lock:
                self.anghaLeaves_passed += 1


    # 编译并运行anghaMath程序组
    def compile_and_run_anghaMath(self,source_file):
        complete_source_file = self.anghaMath / source_file
        log_file = f"{self.anghaMath_logs}/{source_file}.log"
        compile = subprocess.run(
            f"gcc {complete_source_file} -o {self.anghaMath_output}/{source_file}.out",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
        )
        if compile.returncode != 0:
            with self.anghaMath_lock:
                self.anghaMath_failed += 1
                self.ws_anghaMath.append([source_file, 'compile failed',log_file])
            with open(log_file,'w') as log:
                log.write(compile.stdout.decode('utf-8'))
            return

        run = subprocess.run(
            f"{self.anghaMath_output}/{source_file}.out 0",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
        )
        if run.returncode != 0:
            with self.anghaMath_lock:
                self.anghaMath_failed += 1
                self.ws_anghaMath.append([source_file, 'run failed', log_file])
            with open(log_file, 'w') as log:
                log.write(compile.stdout.decode('utf-8') + '\n' + run.stdout.decode('utf-8'))
        else:
            with self.anghaMath_lock:
                self.anghaMath_passed += 1


    def run_test(self):
        anghaLeaves_source_files = list(os.walk(self.anghaLeaves))[0][2]
        anghaMath_source_files = list(os.walk(self.anghaMath))[0][2]
        with ThreadPoolExecutor() as pool:
            pool.map(self.compile_and_run_anghaLeaves, anghaLeaves_source_files)
            pool.map(self.compile_and_run_anghaMath, anghaMath_source_files)

        self.ws_anghaLeaves.append([f'passwd数量:{self.anghaLeaves_passed}',f'failed数量:{self.anghaLeaves_failed}'])
        self.ws_anghaMath.append([f'passwd数量:{self.anghaMath_passed}',f'failed数量:{self.anghaMath_failed}'])
        self.wb.save(self.directory / 'jotai.xlsx')


    def run(self):
        print('开始进行Jotai Benchmark测试')
        self.pre_test()
        self.run_test()
        print('Jotai Benchmark测试结束')