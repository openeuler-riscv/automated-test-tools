import shutil
import subprocess
import os
import asyncio
import numpy
from pathlib import Path
from openpyxl import Workbook
from tqdm import tqdm
from tqdm.asyncio import tqdm as tqdm_asyncio

from .errors import DefaultError


class GpgCheck:
    def __init__(self, **kwargs):
        self.rpms = set()
        self.path = Path('/root/osmts_tmp/gpgcheck')
        self.directory: Path = kwargs.get('saved_directory') / 'gpgcheck'
        self.packages = []

        # 创建Excel表格
        self.wb = Workbook()
        self.ws = self.wb.active


    async def rpm_check_each(self,package_name):
        rpm_check = await asyncio.create_subprocess_shell(
            f"rpm -K /root/osmts_tmp/gpgcheck/{package_name}",
            stdout=asyncio.subprocess.DEVNULL,
            stderr=asyncio.subprocess.PIPE,
        )
        _, stderr = await rpm_check.communicate()
        if rpm_check.returncode != 0: # rpm -K运行失败
            self.ws.append([package_name, stderr.decode('utf-8')])


    async def rpm_check_all(self):
        packages = list(os.walk(self.path))[0][2]
        # 对每个rpm包创建一个测试任务
        tasks = [asyncio.create_task(self.rpm_check_each(package)) for package in packages]
        await tqdm_asyncio.gather(*tasks,leave=False,desc='rpm check')


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir()
        if self.path.exists():
            shutil.rmtree(self.path,ignore_errors=True)
        self.path.mkdir(parents=True)

        # 更新缓存以便后面下载
        try:
            subprocess.run(
                "dnf clean all && dnf makecache",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"gpgcheck测试出错.创建repo缓存失败/")

        # 引入openEuler的gpg验证密钥
        try:
            subprocess.run(
                "rpm --import /etc/pki/rpm-gpg/RPM-GPG-KEY-openEuler",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"gpgcheck测试出错.import gpg文件失败,报错信息:{e.stderr.decode('utf-8')}")


        # 获取已安装的所有rpm包名
        try:
            dnf_list = subprocess.run(
                "dnf list available | awk '/Available Packages/{flag=1; next} flag' | awk '{print $1}'",
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"gpgcheck测试出错.获取所有已安装的rpm包名失败,报错信息:{e.stderr.decode('utf-8')}")
        else:
            self.rpm_package_list = dnf_list.stdout.decode('utf-8').splitlines()

        self.ws.title = 'gpgcheck'
        self.ws.cell(1,1,f"当前系统已安装rpm包的数量:{len(self.rpm_package_list)}")
        self.ws.merge_cells('A1:B1')
        self.ws.cell(2,1,"gpgcheck失败的rpm包统计")
        self.ws.merge_cells('A2:B2')
        self.ws.cell(3,1,"package name")
        self.ws.cell(3,2,"报错日志")


    def run_test(self):
        # 排除掉不符合测试要求的包名
        for package in self.rpm_package_list:
            if package.endswith('.src') or 'debug' in package:
                continue
            else:
                self.packages.append(package)

        print(f"  当前线程的event loop策略:{asyncio.get_event_loop_policy()}")
        # 根据包名批量下载并测试rpm包
        # 分批次原因: shell无法解析长度过大的文本
        piece = int(len(self.packages) / 100)
        for package_list in tqdm(numpy.array_split(self.packages,indices_or_sections=piece),desc='处理包进度',unit='次'):
            rpm_download = subprocess.run(
                f"dnf download {' '.join(package_list)} --destdir={self.path}",
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
            if rpm_download.returncode != 0:
                print(f"gpgcheck测试出错.下载待测的rpm包失败,报错信息:{rpm_download.stderr.decode('utf-8')}")
                print(f"本批次下载出错的rpm包为:{package_list}")

            asyncio.run(self.rpm_check_all())
            shutil.rmtree(self.path)
            self.path.mkdir(parents=True)
        self.wb.save(self.directory / 'gpgcheck.xlsx')


    def run(self):
        print('开始进行gpgcheck测试')
        self.pre_test()
        self.run_test()
        print('gpgcheck测试结束')