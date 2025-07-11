import subprocess,re
from pathlib import Path
from openpyxl import Workbook

from .errors import RunError,SummaryError


class Nmap:
    def __init__(self, **kwargs):
        self.rpms = {'nmap'}
        self.directory:Path = kwargs.get('saved_directory') / 'nmap'
        self.test_result = ''


    def run_test(self):
        try:
            nmap = subprocess.run(
                "nmap -sS -sU 127.0.0.1",
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,e.stderr.decode('utf-8'))
        else:
            self.test_result = nmap.stdout.decode('utf-8')


    def result2summary(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "nmap"
        ws.cell(1,1,"PORT")
        ws.cell(1,2,"STATE")
        ws.cell(1,3,"SERVICE")
        index = 2
        for port,protocol,state,service in re.findall(r"(\d+)\/(tcp|udp)\s+([\w\|\-]+)\s+(\S+)", self.test_result, re.MULTILINE | re.IGNORECASE):
            ws.cell(index,1,port+'/'+protocol)
            ws.cell(index,2,state)
            ws.cell(index,3,service)
            index += 1

        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)
        wb.save(self.directory / 'nmap.xlsx')



    def run(self):
        print("开始进行nmap测试")
        self.run_test()
        try:
            self.result2summary()
        except Exception as e:
            lofFile = self.directory / 'nmap_summary_error.log'
            with open(lofFile,'w') as log:
                log.write(str(e))
            raise SummaryError(lofFile)
        print("nmap测试结束")