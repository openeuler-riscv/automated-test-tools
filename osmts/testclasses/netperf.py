from pathlib import Path
import re,subprocess,psutil
from openpyxl import Workbook
import paramiko
from .errors import DefaultError,RunError


def get_client(ip, password, port=22):
    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy)
    try:
        client.connect(hostname=ip, port=port, username="root", password=password, timeout=100)
    except (
            paramiko.ssh_exception.NoValidConnectionsError,
            paramiko.ssh_exception.AuthenticationException,
            paramiko.ssh_exception.SSHException,
            TypeError,
            AttributeError,
    ) as e:
        print(f"无法连接到远程机器:{ip}.\n原因： {e}")
    return client



class Netperf(object):
    def __init__(self,**kwargs):
        self.rpms = {'netperf'}
        self.directory:Path = kwargs.get('saved_directory') / 'netperf'
        self.server_ip:str = kwargs.get('netperf_server_ip')
        self.netserver_created_by_osmts:bool = kwargs.get('netserver_created_by_osmts')
        self.netserver_created_by_osmts_remote:bool = False
        self.client = None
        self.netperf_server_password = kwargs.get('netperf_server_password')


    def pre_test(self):
        # 非本地测试才会进入pre_test
        self.client = get_client(self.server_ip, self.netperf_server_password,22)
        stdin,stdout,stderr = self.client.exec_command('ps aux|grep netserver|grep -v grep')
        if stdout.channel.recv_exit_status() != 0:
            stdin,stdout,stderr = self.client.exec_command('dnf install netperf-y && netserver -p 10000')
            if stdout.channel.recv_exit_status() != 0:
                raise DefaultError("在远程机器上自动运行netserver失败")
            self.netserver_created_by_osmts_remote = True
        stdin,stdout,stderr = self.client.exec_command('systemctl is-active firewalld')
        self.is_active = stdout.read()
        if self.is_active == "active":
            stdin, stdout, stderr = self.client.exec_command("systemctl stop firewalld")


    def run_test(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "netperf"

        ws.cell(1,1,"TCP STREAM TEST")
        ws.merge_cells("a1:a5")

        ws.cell(6,1,"UDP STREAM TEST")
        ws.merge_cells("a6:a18")

        ws.cell(20,1,"TCP REQUEST/RESPONSE TEST")
        ws.merge_cells("a20:a21")

        ws.cell(22,1,"TCP Connect/Request/Response TEST")
        ws.merge_cells("a22:a23")

        ws.cell(24,1,"UDP REQUEST/RESPONSE TEST")
        ws.merge_cells("a24:a25")

        # TCP_STREAM表头
        ws.cell(1,2,"Recv Socket Size bytes")
        ws.cell(1,3,"Send Socket Size Bytes")
        ws.cell(1,4,"Send Message Size Bytes")
        ws.cell(1,5,"Elapsed Time secs.")
        ws.cell(1,6,"Throughput(10^6bits/sec)")

        # UDP_STREAM表头
        ws.cell(6,2,"Socket Size bytes")
        ws.cell(6,3,"Message Size bytes")
        ws.cell(6,4,"Elapsed Time secs")
        ws.cell(6,5,"Messages Okay")
        ws.cell(6, 6, "Messages Errors")
        ws.cell(6,7,"Throughput(10^6bits/sec)")

        # 剩余三个测试的表头
        ws.cell(19,2,"Local Socket Send bytes")
        ws.cell(19, 3, "Remote Size Recv Bytes")
        ws.cell(19, 4, "Request Size bytes")
        ws.cell(19, 5, "Resp. Size bytes")
        ws.cell(19, 6, "Elapsed Time secs.")
        ws.cell(19,7,"Trans. Rate per sec")




        # TCP_STREAM测试
        line = 2
        for message_size_bytes in (1,64,512,65536):
            try:
                TCP_STREAM = subprocess.run(
                    f"netperf -t TCP_STREAM -H {self.server_ip} -p 10000 -l 60 -- -m {message_size_bytes}",
                    shell=True,check=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise RunError(e.returncode,f"netperf测试出错:TCP_STREAM测试运行失败.报错信息:{e.stderr}")
            result = re.findall(r'\d+\.\d+|\d+', TCP_STREAM.stdout.decode('utf-8').split('\n')[6])
            for col,value in enumerate(result):
                ws.cell(line,col+2,value)
            line += 1


        # UDP_STREAM测试
        line = 7
        for message_size_bytes in (1,64,128,256,512,32768):
            try:
                UDP_STREAM = subprocess.run(
                    f"netperf -t UDP_STREAM -H {self.server_ip} -p 10000 -l 60 -- -m {message_size_bytes}",
                    shell=True,check=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise RunError(e.returncode,f"netperf测试出错:UDP_STREAM测试运行失败.报错信息:{e.stderr.decode('utf-8')}")
            result1 = re.findall(r'\d+\.\d+|\d+', UDP_STREAM.stdout.decode('utf-8').split('\n')[5])
            result2 = re.findall(r'\d+\.\d+|\d+', UDP_STREAM.stdout.decode('utf-8').split('\n')[6])
            for col,value in enumerate(result1):
                ws.cell(line,col+2,value)
            line += 1
            ws.cell(line,2,result2[0])
            ws.cell(line, 4, result2[1])
            ws.cell(line, 5, result2[2])
            ws.cell(line, 7, result2[3])
            line += 1


        # TCP REQUEST/RESPONSE测试
        try:
            TCP_RR = subprocess.run(
                f"netperf -t TCP_RR -H {self.server_ip} -p 10000",
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,f"netperf测试出错:TCP_RR测试运行失败.报错信息:{e.stderr.decode('utf-8')}")
        result1 = re.findall(r'\d+\.\d+|\d+', TCP_RR.stdout.decode('utf-8').split('\n')[6])
        result2 = re.findall(r'\d+\.\d+|\d+', TCP_RR.stdout.decode('utf-8').split('\n')[7])
        for col, value in enumerate(result1):
            ws.cell(20, col + 2, value)
        ws.cell(21, 2, result2[0])
        ws.cell(21, 3, result2[1])


        # TCP Connect/Request/Response测试
        try:
            TCP_CRR = subprocess.run(
                f"netperf -t TCP_CRR -H {self.server_ip} -p 10000",
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,f"netperf测试出错:TCP_CRR测试运行失败.报错信息:{e.stderr.decode('utf-8')}")
        result1 = re.findall(r'\d+\.\d+|\d+', TCP_CRR.stdout.decode('utf-8').split('\n')[6])
        result2 = re.findall(r'\d+\.\d+|\d+', TCP_CRR.stdout.decode('utf-8').split('\n')[7])
        for col, value in enumerate(result1):
            ws.cell(22, col + 2, value)
        ws.cell(23, 2, result2[0])
        ws.cell(23, 3, result2[1])


        # UDP REQUEST/RESPONSE测试
        try:
            UDP_RR = subprocess.run(
                f"netperf -t UDP_RR -H {self.server_ip} -p 10000",
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,f"netperf测试出错:UDP_RR测试运行失败.报错信息:{UDP_RR.stderr.decode('utf-8')}")
        result1 = re.findall(r'\d+\.\d+|\d+', UDP_RR.stdout.decode('utf-8').split('\n')[6])
        result2 = re.findall(r'\d+\.\d+|\d+', UDP_RR.stdout.decode('utf-8').split('\n')[7])
        for col, value in enumerate(result1):
            ws.cell(24, col + 2, value)
        ws.cell(25, 2, result2[0])
        ws.cell(25, 3, result2[1])

        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)
        wb.save(self.directory / 'netperf.xlsx')


    def post_test(self):
        if self.netserver_created_by_osmts:
            for process in psutil.process_iter():
                if process.name() == 'netserver':
                    process.terminate()
        if self.client is not None:
            if self.netserver_created_by_osmts_remote:
                self.client.exec_command("pkill -9 netserver")
            if self.is_active == 'active':
                self.client.exec_command("systemctl start firewalld")


    def run(self):
        print("开始进行netperf测试")
        if self.netperf_server_password is not None:
            self.pre_test()
        self.run_test()
        self.post_test()
        print("netperf测试结束")