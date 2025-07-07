import os
import tempfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Alignment
from config import config

def convert_xls_to_xlsx(xls_path):
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        xls = pd.ExcelFile(xls_path, engine="xlrd")
    except Exception as e:
        raise RuntimeError(f"无法读取 .xls 文件：{xls_path}\n请确认是否是有效 Excel 文件，且安装了 xlrd（支持 .xls）\n原始错误：{e}")
    with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, engine="xlrd")
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    return tmp_path

def load_workbook_auto(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        tmp_xlsx = convert_xls_to_xlsx(file_path)
        wb = load_workbook(tmp_xlsx, data_only=True)
        os.remove(tmp_xlsx)
        return wb
    else:
        return load_workbook(file_path, data_only=True)

def read_range(ws, cell_range, transpose=False):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    data = []
    print(f"[DEBUG] Read range {cell_range}, transpose={transpose}")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        print(f"[DEBUG] Processing row: {[cell.value for cell in row]}")
        row_data = []
        for cell in row:
            val = cell.value
            if isinstance(val, str):
                try:
                    val = float(val) if '.' in val else int(val)
                except:
                    pass
            row_data.append(val)
        data.append(row_data)
    if transpose:
        data = list(map(list, zip(*data)))
    return data

def write_data(ws, target_range, data):
    min_col, min_row, _, _ = range_boundaries(target_range)
    for r_idx, row_data in enumerate(data):
        for c_idx, val in enumerate(row_data):
            ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=val)

def copy_multi_ranges(ws_src, ws_dst, src_ranges, dst_ranges, transpose_flags=None):
    if transpose_flags is None:
        transpose_flags = [False] * len(src_ranges)
    for src_range, dst_range, transpose in zip(src_ranges, dst_ranges, transpose_flags):
        data = read_range(ws_src, src_range, transpose)
        write_data(ws_dst, dst_range, data)

def copy_test_type_project(ws_src, ws_dst, configs):
    for cfg in configs:
        # 新版：单个 source 对多个 target
        if "source_range" in cfg:
            src_range = cfg["source_range"]
            tgt_ranges = cfg["target_ranges"]
            transpose_flags = cfg.get("transpose", [False] * len(tgt_ranges))
            for tgt_range, transpose in zip(tgt_ranges, transpose_flags):
                data = read_range(ws_src, src_range, transpose)
                write_data(ws_dst, tgt_range, data)
        # 兼容旧版：多个 source 对应多个 target
        elif "source_ranges" in cfg:
            src_ranges = cfg["source_ranges"]
            tgt_ranges = cfg["target_ranges"]
            transpose_flags = cfg.get("transpose", [False] * len(tgt_ranges))
            for src_range, tgt_range, transpose in zip(src_ranges, tgt_ranges, transpose_flags):
                data = read_range(ws_src, src_range, transpose)
                write_data(ws_dst, tgt_range, data)

def generate_diff_formula(ws, target_range, ver1_ranges, ver2_ranges, formula_template):
    min_col_t, min_row_t, max_col_t, max_row_t = range_boundaries(target_range)
    rows = max_row_t - min_row_t + 1
    cols = max_col_t - min_col_t + 1
    min_col_v1, min_row_v1, _, _ = range_boundaries(ver1_ranges[0])
    min_col_v2, min_row_v2, _, _ = range_boundaries(ver2_ranges[0])
    for r in range(rows):
        for c in range(cols):
            cell = ws.cell(row=min_row_t + r, column=min_col_t + c)
            ver1_cell_addr = f"{get_column_letter(min_col_v1 + c)}{min_row_v1 + r}"
            ver2_cell_addr = f"{get_column_letter(min_col_v2 + c)}{min_row_v2 + r}"

            ver1_val = ws[ver1_cell_addr].value
            ver2_val = ws[ver2_cell_addr].value

            if ver1_val in (None, "") or ver2_val in (None, ""):
                cell.value = None
            else:
                formula = formula_template.format(ver1=ver1_cell_addr, ver2=ver2_cell_addr)
                cell.value = formula
                cell.alignment = Alignment(horizontal='center')

def main():
    versions = config["versions"]
    devices = config["devices"]
    for device in devices:
        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        for tool_name, tool_conf in config["test_tools"].items():
            ws_out = wb_out.create_sheet(title=tool_name)
            file_ver1 = os.path.join(versions[0], device, tool_name, tool_conf["filename"])
            file_ver2 = os.path.join(versions[1], device, tool_name, tool_conf["filename"])
            if not (os.path.exists(file_ver1) and os.path.exists(file_ver2)):
                print(f"缺少文件：{file_ver1} 或 {file_ver2}，跳过 {device} - {tool_name}")
                continue
            wb1 = load_workbook_auto(file_ver1)
            wb2 = load_workbook_auto(file_ver2)
            ws1 = wb1.active
            ws2 = wb2.active
            vm = tool_conf["version_mark"]
            ws_out[vm["result_cells"][0]].value = vm["ver1_text"]
            ws_out[vm["result_cells"][1]].value = vm["ver2_text"]
            diff_conf = tool_conf.get("diff_result", {})
            if "difference_title_cell" in diff_conf:
                ws_out[diff_conf["difference_title_cell"]] = diff_conf["difference_title_text"]
            copy_test_type_project(ws1, ws_out, tool_conf["test_type_project"])
            ver1 = tool_conf["result_data"]["ver1"]
            ver2 = tool_conf["result_data"]["ver2"]
            copy_multi_ranges(ws1, ws_out, ver1["source_ranges"], ver1["target_ranges"], [ver1.get("transpose", False)] * len(ver1["source_ranges"]))
            copy_multi_ranges(ws2, ws_out, ver2["source_ranges"], ver2["target_ranges"], [ver2.get("transpose", False)] * len(ver2["source_ranges"]))
            generate_diff_formula(ws_out, diff_conf.get("target_range", ""), ver1["target_ranges"], ver2["target_ranges"], diff_conf.get("formula", ""))
        out_file = f"comparison_{device}.xlsx"
        wb_out.save(out_file)
        print(f"设备 {device} 对比文件生成：{out_file}")

if __name__ == "__main__":
    main()
